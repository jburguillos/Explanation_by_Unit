import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

# Leer archivo Excel
def read_excel_file(file_path, sheet_name):
    try:
        with open(file_path, 'rb') as f:
            df = pd.read_excel(f, sheet_name=sheet_name, engine='openpyxl')
        df.columns = [col.strip('´').strip() for col in df.columns]
        return df
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return None

# Extraer mercados únicos
def extract_unique_markets(df, market_column):
    return df[market_column].unique() if df is not None else []

# Año fiscal y trimestre
def get_financial_year_and_quarter():
    now = datetime.now()
    year = now.year
    month = now.month
    financial_year = year if month >= 4 else year - 1
    quarter = ['Q4', 'Q1', 'Q1', 'Q1', 'Q2', 'Q2', 'Q2', 'Q3', 'Q3', 'Q3', 'Q4', 'Q4'][month - 1]
    return financial_year, quarter

# Marcar contratos top
def mark_top_contracts(df, column):
    total_change = df[column].abs().sum()
    df_sorted = df.reindex(df[column].abs().sort_values(ascending=False).index)
    cumulative_sum = df_sorted[column].abs().cumsum()
    df_sorted['Cumulative Contribution'] = cumulative_sum
    df_sorted['Cumulative Percentage'] = (cumulative_sum / total_change) * 100
    df_sorted['TBC'] = ''
    df_sorted.loc[cumulative_sum <= 0.8 * total_change, 'TBC'] = 'X'
    df_sorted.loc[df_sorted[column].abs() > 200000, 'TBC'] = 'XX'
    return df_sorted.sort_index()

# Asignar efectos
def assign_effects(row):
    effects = []
    if pd.notna(row['TUs']) and isinstance(row['TUs'], (int, float)) and row['TUs'] != 0:
        effects.append(('True Up (TU)', row['TUs']))
    if pd.notna(row['Change Prognosis']) and isinstance(row['Change Prognosis'], (int, float)) and row['Change Prognosis'] != 0:
        effects.append(('Change in Prognosis (∆Prog)', row['Change Prognosis']))
    
    # Filtrar efectos con valores numéricos válidos
    effects = [e for e in effects if isinstance(e[1], (int, float))]
    
    # Ordenar por impacto absoluto
    effects.sort(key=lambda x: abs(x[1]), reverse=True)

    effect_data = {}
    for i, (effect, impact) in enumerate(effects[:4], start=1):
        effect_data[f'Effect {i}'] = effect
        effect_data[f'Impact {i} (m€)'] = round(impact / 1_000_000, 2)
        effect_data[f'Type {i}'] = ''
    return pd.Series(effect_data)


# Formato general
def apply_formatting(ws):
    for cell in ws["1:1"]:
        cell.font = Font(bold=True)
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
    fill_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    fill_even = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.fill = fill_even if cell.row % 2 == 0 else fill_odd
    ws.freeze_panes = ws["A2"]

# Formato condicional
def apply_conditional_formatting(ws):
    for col in ws.iter_cols(1, ws.max_column):
        if col[0].value == 'Check Effect':
            col_letter = get_column_letter(col[0].column)
            rule = ColorScaleRule(
                start_type='num', start_value=-0.035, start_color='FF0000',
                mid_type='num', mid_value=0, mid_color='00FF00',
                end_type='num', end_value=0.035, end_color='FF0000'
            )
            ws.conditional_formatting.add(f"{col_letter}2:{col_letter}{ws.max_row}", rule)
            break

# Crear hojas por mercado
def create_market_sheets(df, unique_markets, market_column, output_file):
    if df is None or len(unique_markets) == 0:
        print("Error: Unable to create market sheets")
        return

    filtered_df = df[df['Entry Code Accounting Principle TD'] == 'Result']
    financial_year, financial_quarter = get_financial_year_and_quarter()

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        summary_data = []

        for market in unique_markets:
            market_df = filtered_df[filtered_df[market_column] == market].copy()
            market_df = mark_top_contracts(market_df, 'Change in HFC Q(2.2025 –1.2025) in OPT1')

            new_df = pd.DataFrame({
                'Technical Result Previous Quarter (m€)': market_df['HFC as at Q1.2025 in OPT1'].apply(lambda x: round(x / 1_000_000, 2)),
                'Technical Result Current Quarter (m€)': market_df['HFC as at Q2.2025 in OPT1'].apply(lambda x: round(x / 1_000_000, 2)),
                'Total impact (m€)': market_df['Change in HFC Q(2.2025 –1.2025) in OPT1'].apply(lambda x: round(x / 1_000_000, 2)),
                'Check Effect': '',
                'Financial Year': financial_year,
                'Financial Quarter': financial_quarter,
                'CAP': '',
                'Class of Business': market_df['Class of Business'],
                'Market': market_df['Business Unit'],
                'BUPA Name': market_df['BUPA'],
                'Treaty Number': market_df['Contract TD'],
                'Treaty Name': market_df['Contract Name'],
                'CoB': '',
                'Effect 1': '',
                'Type 1': '',
                'Impact 1 (m€)': '',
                'Effect 2': '',
                'Type 2': '',
                'Impact 2 (m€)': '',
                'Effect 3': '',
                'Type 3': '',
                'Impact 3 (m€)': '',
                'Effect 4': '',
                'Type 4': '',
                'Impact 4 (m€)': '',
                'Comments': '',
                'TBC': market_df['TBC']
            })

            effect_df = market_df.apply(assign_effects, axis=1)
            for col in effect_df.columns:
                new_df[col] = effect_df[col]

            for i in range(len(new_df)):
                new_df.at[new_df.index[i], 'Check Effect'] = f"=C{i+2}-P{i+2}-S{i+2}-V{i+2}-Y{i+2}"

            sheet_name = str(market)[:31]
            new_df.to_excel(writer, sheet_name=sheet_name, index=False)
###  BREAKDOWN sheets & Summary
            breakdown_df = market_df[['Contract TD',  'BUPA', 'Change in HFC Q(2.2025 –1.2025) in OPT1',
                                      'Cumulative Contribution', 'Cumulative Percentage', 'TBC']]
            breakdown_df.columns = ['Contract ID', 'BUPA Name', 'Absolute HFC Change',
                                    'Cumulative Contribution', 'Cumulative Percentage', 'Top 80% Contributor']
            breakdown_df = breakdown_df.sort_values(by='Absolute HFC Change', ascending=False)
            breakdown_df.to_excel(writer, sheet_name=f"Breakdown - {str(market)[:28]}", index=False)

            summary_data.append({
                'Market': market,
                'Total Contracts': new_df.shape[0],
                'Unique Legal Partners': market_df['Legal Partner TD'].nunique(),
                'Total CNs': market_df['Contract TD'].nunique(),
                'Total HFC Change': new_df['Total impact (m€)'].sum(),
                'Average HFC Change': new_df['Total impact (m€)'].mean(),
                'Average HFC Q2.2025': new_df['Technical Result Current Quarter (m€)'].mean(),
                'Average HFC Q1.2025': new_df['Technical Result Previous Quarter (m€)'].mean(),
                'Positive HFC Changes': new_df[new_df['Total impact (m€)'] > 0].shape[0],
                'Negative HFC Changes': new_df[new_df['Total impact (m€)'] < 0].shape[0]
            })
###
        pd.DataFrame(summary_data).to_excel(writer, sheet_name='Summary', index=False)
        filtered_df.nlargest(250, 'Change in HFC Q(2.2025 –1.2025) in OPT1').to_excel(writer, sheet_name='Top Positive HFC Changes', index=False)
        filtered_df.nsmallest(250, 'Change in HFC Q(2.2025 –1.2025) in OPT1').to_excel(writer, sheet_name='Top Negative HFC Changes', index=False)

    wb = load_workbook(output_file)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        apply_formatting(ws)
        apply_conditional_formatting(ws)
        ws.auto_filter.ref = ws.dimensions #### working?? YYEEEEEAAAA
    wb.save(output_file)

# USO DE EJEMPLO
file_path = "Q3_input.xlsx"
sheet_name = "DBQ2"
market_column = "Business Unit"
output_file = "Outputtest2.xlsx"

df = read_excel_file(file_path, sheet_name)
unique_markets = extract_unique_markets(df, market_column)
create_market_sheets(df, unique_markets, market_column, output_file)

